# stok-app.py — Admin Menü + Ürün Yönetimi (taslak/yayın + Excel) + Kullanıcı Yönetimi + Kampanya Pop-up (yükle/sil) + Bayi vitrini
# Tarih: 10.09.2025
# Bu sürümde:
# - Ürün kodu şeritleri (dikey) hem admin hem bayi tarafında hücrede TAM ORTALI (flex).
# - Kod sütunu başlıkları merkezli.
# - Container genişlikleri: 1600px.
# - Diğer fonksiyonlar korunmuştur (Excel, taslak/yayın, kampanya pop-up, kullanıcılar).

from fastapi import FastAPI, Request, Form, UploadFile, File, HTTPException
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.middleware.sessions import SessionMiddleware
from pydantic import BaseModel
from typing import List, Optional
import sqlite3, os, secrets, io
from datetime import datetime
from openpyxl import load_workbook

APP_TITLE = "Canlı Stok Portalı"
CENTER_LOCATION_CODE = os.environ.get("CENTER_CODE", "MERKEZ")

STATIC_DIR = "static"
UPLOAD_DIR = os.path.join(STATIC_DIR, "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

app = FastAPI(title=APP_TITLE)
app.add_middleware(SessionMiddleware, secret_key=os.environ.get("SESSION_SECRET","super-secret-key-please-change"))
templates = Jinja2Templates(directory="templates_inline")
app.mount("/static", StaticFiles(directory=STATIC_DIR), name="static")

# Kampanya kategorileri
CAMPAIGN_CATEGORIES = [
    "DuraLife Collection",
    "DuraLife Mobilya",
    "Batarya",
    "Fırsat Akrilik",
]
# Ürün Kategorileri
PRODUCT_CATEGORIES = [
    "Vitrifiye",
    "Mobilya",
    "Lavabo Bataryası",
    "Banyo Bataryası",
    "Küvet",
]

DB_PATH = os.environ.get("DB_PATH", "stock.db")
def db():
    con = sqlite3.connect(DB_PATH)
    con.row_factory = sqlite3.Row
    return con

def init_db():
    con = db(); c = con.cursor()
    # users
    c.execute("""
    CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        is_active INTEGER DEFAULT 1,
        created_at TEXT
    )""")
    # products
    c.execute("""
    CREATE TABLE IF NOT EXISTS product(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT UNIQUE,
        description TEXT DEFAULT '',
        image_path TEXT DEFAULT '',
        list_price REAL DEFAULT 0,
        sale_price REAL DEFAULT 0,
        cargo_fee TEXT DEFAULT '0',
        durapay REAL DEFAULT 0,
        campaign_categories TEXT DEFAULT '',
        product_category TEXT DEFAULT 'Vitrifiye',
        is_active INTEGER DEFAULT 1,
        created_at TEXT
    )""")
    # locations
    c.execute("""CREATE TABLE IF NOT EXISTS location(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT, code TEXT UNIQUE
    )""")
    # stock
    c.execute("""CREATE TABLE IF NOT EXISTS stock_snapshot(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        product_id INTEGER, location_id INTEGER, onhand REAL DEFAULT 0,
        updated_at TEXT,
        UNIQUE(product_id, location_id)
    )""")
    # campaign popups
    c.execute("""
    CREATE TABLE IF NOT EXISTS campaign_popup(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        image_path TEXT NOT NULL,
        is_active INTEGER DEFAULT 1,
        sort_order INTEGER DEFAULT 0,
        created_at TEXT
    )""")
    # migrations for product
    cols = {r[1] for r in c.execute("PRAGMA table_info(product)").fetchall()}
    if "cargo_fee" not in cols: c.execute("ALTER TABLE product ADD COLUMN cargo_fee TEXT DEFAULT '0'")
    if "list_price" not in cols: c.execute("ALTER TABLE product ADD COLUMN list_price REAL DEFAULT 0")
    if "sale_price" not in cols: c.execute("ALTER TABLE product ADD COLUMN sale_price REAL DEFAULT 0")
    if "description" not in cols: c.execute("ALTER TABLE product ADD COLUMN description TEXT DEFAULT ''")
    if "image_path" not in cols: c.execute("ALTER TABLE product ADD COLUMN image_path TEXT DEFAULT ''")
    if "created_at" not in cols: c.execute("ALTER TABLE product ADD COLUMN created_at TEXT")
    if "durapay" not in cols: c.execute("ALTER TABLE product ADD COLUMN durapay REAL DEFAULT 0")
    if "campaign_categories" not in cols: c.execute("ALTER TABLE product ADD COLUMN campaign_categories TEXT DEFAULT ''")
    if "product_category" not in cols: c.execute("ALTER TABLE product ADD COLUMN product_category TEXT DEFAULT 'Vitrifiye'")
    if "category" in cols:
        c.execute("UPDATE product SET campaign_categories=CASE WHEN IFNULL(campaign_categories,'')='' THEN category ELSE campaign_categories END")
    # default location
    c.execute("SELECT id FROM location WHERE code=?", (CENTER_LOCATION_CODE,))
    if not c.fetchone():
        c.execute("INSERT INTO location(name,code) VALUES(?,?)", ("Ana Merkez", CENTER_LOCATION_CODE))
    # seed admin user
    c.execute("SELECT id FROM users WHERE username='admin'")
    if not c.fetchone():
        c.execute("INSERT INTO users(username,password,is_active,created_at) VALUES(?,?,1,?)",
                  ("admin","admin123", datetime.utcnow().isoformat(timespec="seconds")))
    con.commit(); con.close()

def get_location_id(code: str) -> int:
    con=db(); c=con.cursor()
    c.execute("SELECT id FROM location WHERE code=?", (code,))
    row=c.fetchone()
    if not row:
        c.execute("INSERT INTO location(name,code) VALUES(?,?)", (code, code))
        con.commit()
        c.execute("SELECT id FROM location WHERE code=?", (code,))
        row=c.fetchone()
    loc_id=row["id"]; con.close(); return loc_id

def get_onhand(product_id:int, location_code:str)->float:
    con=db(); c=con.cursor()
    loc_id=get_location_id(location_code)
    c.execute("SELECT onhand FROM stock_snapshot WHERE product_id=? AND location_id=?",(product_id,loc_id))
    row=c.fetchone(); con.close()
    return float(row["onhand"]) if row else 0.0

def set_snapshot(product_id:int, location_code:str, onhand:float):
    con=db(); c=con.cursor()
    loc_id=get_location_id(location_code)
    c.execute("SELECT id FROM stock_snapshot WHERE product_id=? AND location_id=?", (product_id, loc_id))
    row = c.fetchone()
    now = datetime.utcnow().isoformat(timespec="seconds")
    if row:
        c.execute("UPDATE stock_snapshot SET onhand=?, updated_at=? WHERE id=?", (onhand, now, row["id"]))
    else:
        c.execute("""INSERT INTO stock_snapshot(product_id,location_id,onhand,updated_at)
                     VALUES(?,?,?,?)""", (product_id, loc_id, onhand, now))
    con.commit(); con.close()

def unique_product_name(desired_name: str, exclude_id: int | None = None) -> str:
    base = (desired_name or "").strip() or "Ürün"
    con = db(); c = con.cursor()
    candidate = base; i = 1
    while True:
        if exclude_id is None:
            c.execute("SELECT id FROM product WHERE name=?", (candidate,))
        else:
            c.execute("SELECT id FROM product WHERE name=? AND id<>?", (candidate, exclude_id))
        row = c.fetchone()
        if not row:
            con.close(); return candidate
        i += 1; candidate = f"{base} - {i}"

class RedirectException(Exception):
    def __init__(self, url:str): self.url=url
@app.exception_handler(RedirectException)
async def _redir(request:Request, exc:RedirectException): return RedirectResponse(exc.url, status_code=303)
def require_login(request:Request):
    if not request.session.get("user"): raise RedirectException("/login")

@app.on_event("startup")
def _startup():
    os.makedirs("templates_inline", exist_ok=True)
    _materialize_templates()
    init_db()

@app.get("/", include_in_schema=False)
def root(): return RedirectResponse("/login", status_code=303)

@app.get("/health", include_in_schema=False)
def health(): return {"ok": True}

# ---------- AUTH ----------
@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request):
    if request.session.get("user"):
        return RedirectResponse("/admin", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "title": APP_TITLE})

@app.post("/login")
def login_submit(request: Request, username: str = Form(...), password: str = Form(...)):
    con=db(); c=con.cursor()
    c.execute("SELECT id FROM users WHERE username=? AND password=? AND is_active=1", (username, password))
    row = c.fetchone(); con.close()
    if row:
        request.session["user"]=username
        return RedirectResponse("/admin", status_code=303)
    return templates.TemplateResponse("login.html", {"request": request, "title": APP_TITLE, "error":"Hatalı kullanıcı adı veya şifre."})

@app.get("/logout")
def logout(request: Request):
    request.session.clear()
    return RedirectResponse("/login", status_code=303)

# ===================== ADMIN MENÜ =====================
@app.get("/admin", response_class=HTMLResponse)
def admin_menu(request: Request):
    require_login(request)
    return templates.TemplateResponse("admin_menu.html", {
        "request": request, "title": APP_TITLE, "username": request.session.get("user")
    })

# ===================== ÜRÜN YÖNETİMİ =====================
@app.get("/admin/products", response_class=HTMLResponse)
def admin_products(request: Request, up_ok: int = 0, up_new: int = 0, up_err: int = 0):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("""
    SELECT p.*, COALESCE(ss.onhand,0) as onhand
    FROM product p
    LEFT JOIN location l ON l.code=?
    LEFT JOIN stock_snapshot ss ON ss.product_id=p.id AND ss.location_id=l.id
    ORDER BY p.id DESC
    """,(CENTER_LOCATION_CODE,))
    products=c.fetchall(); con.close()
    return templates.TemplateResponse("admin_products.html", {
        "request": request, "products": products,
        "title": APP_TITLE, "username": request.session.get("user"),
        "campaign_cats": CAMPAIGN_CATEGORIES, "product_cats": PRODUCT_CATEGORIES,
        "up_ok": up_ok, "up_new": up_new, "up_err": up_err
    })

@app.post("/admin/product/create")
async def admin_product_create(
    request: Request,
    name: str = Form(...),
    description: str = Form(""),
    list_price: float = Form(0),
    sale_price: float = Form(0),
    durapay: float = Form(0),
    product_category: str = Form(PRODUCT_CATEGORIES[0]),
    campaign_categories: Optional[List[str]] = Form(None),
    stock: float = Form(0),
    cargo_fee: str = Form("0"),
    file: UploadFile = File(None),
    save_mode: str = Form("publish")
):
    require_login(request)
    image_path = ""
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in (".jpg",".jpeg",".png",".webp",".gif"):
            raise HTTPException(400, "Sadece .jpg, .jpeg, .png, .webp, .gif kabul edilir")
        fname = f"{secrets.token_hex(6)}{ext}"
        with open(os.path.join(UPLOAD_DIR, fname), "wb") as f: f.write(await file.read())
        image_path = f"/static/uploads/{fname}"

    final_name = unique_product_name(name.strip())
    cc_list = [c for c in (campaign_categories or []) if c in CAMPAIGN_CATEGORIES]
    cc_csv = "," + ",".join(cc_list) + "," if cc_list else ""
    active_flag = 1 if (save_mode or "publish") == "publish" else 0

    con=db(); c=con.cursor()
    now=datetime.utcnow().isoformat(timespec="seconds")
    c.execute("""INSERT INTO product(name,description,image_path,list_price,sale_price,cargo_fee,durapay,campaign_categories,product_category,is_active,created_at)
                 VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
              (final_name, description.strip(), image_path,
               float(list_price), float(sale_price), cargo_fee.strip(), float(durapay),
               cc_csv, product_category.strip(), active_flag, now))
    pid=c.lastrowid
    con.commit(); con.close()

    set_snapshot(pid, CENTER_LOCATION_CODE, float(stock))
    return RedirectResponse("/admin/products", status_code=303)

@app.get("/admin/product/edit/{pid}", response_class=HTMLResponse)
def product_edit_page(request: Request, pid:int):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("SELECT * FROM product WHERE id=?", (pid,))
    p=c.fetchone()
    if not p: 
        con.close(); raise HTTPException(404, "Ürün bulunamadı")
    current_stock = get_onhand(pid, CENTER_LOCATION_CODE)
    con.close()
    current_cc = set([x for x in (p["campaign_categories"] or "").strip(",").split(",") if x])
    return templates.TemplateResponse("edit.html", {
        "request": request, "p": p, "stock": current_stock,
        "title": APP_TITLE, "campaign_cats": CAMPAIGN_CATEGORIES,
        "product_cats": PRODUCT_CATEGORIES, "current_cc": current_cc
    })

@app.post("/admin/product/update")
async def product_update(
    request: Request,
    pid: int = Form(...),
    name: str = Form(...),
    description: str = Form(""),
    list_price: float = Form(0),
    sale_price: float = Form(0),
    durapay: float = Form(0),
    product_category: str = Form(PRODUCT_CATEGORIES[0]),
    campaign_categories: Optional[List[str]] = Form(None),
    stock: float = Form(0),
    cargo_fee: str = Form("0"),
    file: UploadFile = File(None)
):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("SELECT * FROM product WHERE id=?", (pid,))
    prev=c.fetchone()
    if not prev:
        con.close(); raise HTTPException(404, "Ürün bulunamadı")

    image_path = prev["image_path"] or ""
    if file and file.filename:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in (".jpg",".jpeg",".png",".webp",".gif"):
            con.close(); raise HTTPException(400, "Sadece .jpg, .jpeg, .png, .webp, .gif kabul edilir")
        fname = f"{secrets.token_hex(6)}{ext}"
        with open(os.path.join(UPLOAD_DIR, fname), "wb") as f: f.write(await file.read())
        image_path = f"/static/uploads/{fname}"

    safe_name = unique_product_name(name.strip(), exclude_id=pid)
    cc_list = [c for c in (campaign_categories or []) if c in CAMPAIGN_CATEGORIES]
    cc_csv = "," + ",".join(cc_list) + "," if cc_list else ""

    c.execute("""UPDATE product
                 SET name=?, description=?, image_path=?, list_price=?, sale_price=?, cargo_fee=?, durapay=?, campaign_categories=?, product_category=?, is_active=1
                 WHERE id=?""",
              (safe_name, description.strip(), image_path, float(list_price), float(sale_price),
               cargo_fee.strip(), float(durapay), cc_csv, product_category.strip(), pid))
    con.commit(); con.close()

    set_snapshot(pid, CENTER_LOCATION_CODE, float(stock))
    return RedirectResponse("/admin/products", status_code=303)

# ===================== EXCEL YÜKLEME =====================
@app.post("/admin/products/upload-excel")
async def upload_excel(request: Request, xls: UploadFile = File(...)):
    require_login(request)
    if not xls or not xls.filename:
        raise HTTPException(400, "Excel dosyası seçilmedi.")
    ext = os.path.splitext(xls.filename)[1].lower()
    if ext not in (".xlsx",):
        raise HTTPException(400, "Lütfen .xlsx (Excel) dosyası yükleyin.")

    content = await xls.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Excel okunamadı: {e}")

    ws = wb.active
    headers = {}
    for cell in ws[1]:
        if cell.value is None: continue
        headers[str(cell.value).strip().lower()] = cell.column

    def find_col(names):
        for n in names:
            if n.lower() in headers: return headers[n.lower()]
        return None

    c_item = find_col(["Item"])
    c_qty  = find_col(["Available Qnt", "Avaliable Qnt"])  # eski yanlış yazım da kabul

    if not c_item or not c_qty:
        raise HTTPException(400, "Gerekli başlıklar bulunamadı: 'Item' ve 'Available Qnt'.")

    up_ok = up_new = up_err = 0
    con=db(); c=con.cursor()
    now = datetime.utcnow().isoformat(timespec="seconds")

    for r in ws.iter_rows(min_row=2):
        try:
            item = r[c_item-1].value
            qty  = r[c_qty-1].value
            if item is None: up_err+=1; continue
            name = str(item).strip()
            if not name: up_err+=1; continue
            try: q = float(str(qty).replace(",", ".")) 
            except: q = 0.0

            c.execute("SELECT id FROM product WHERE name=?", (name,))
            row = c.fetchone()
            if row:
                set_snapshot(row["id"], CENTER_LOCATION_CODE, q)
                up_ok += 1
            else:
                c.execute("""INSERT INTO product(name, description, image_path, list_price, sale_price, cargo_fee, durapay, campaign_categories, product_category, is_active, created_at)
                             VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                          (name, "", "", 0.0, 0.0, "0", 0.0, "", PRODUCT_CATEGORIES[0], 0, now))
                con.commit()
                set_snapshot(c.lastrowid, CENTER_LOCATION_CODE, q)
                up_new += 1
        except Exception:
            up_err += 1
            continue
    con.close()
    return RedirectResponse(f"/admin/products?up_ok={up_ok}&up_new={up_new}&up_err={up_err}", status_code=303)

# ===================== KAMPANYA POP-UP =====================
@app.get("/admin/campaigns", response_class=HTMLResponse)
def admin_campaigns(request: Request):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("SELECT * FROM campaign_popup WHERE is_active=1 ORDER BY sort_order ASC, id DESC")
    popups = c.fetchall(); con.close()
    return templates.TemplateResponse("admin_campaigns.html", {
        "request": request, "title": APP_TITLE, "username": request.session.get("user"), "popups": popups
    })

@app.post("/admin/campaign/upload")
async def admin_campaign_upload(request: Request, files: List[UploadFile] = File(...)):
    require_login(request)
    if not files:
        raise HTTPException(400, "En az bir görsel seçin.")
    con=db(); c=con.cursor()
    now=datetime.utcnow().isoformat(timespec="seconds")
    for file in files:
      if not file.filename: 
          continue
      ext = os.path.splitext(file.filename)[1].lower()
      if ext not in (".jpg",".jpeg",".png",".webp",".gif"):
          continue
      fname = f"camp_{secrets.token_hex(6)}{ext}"
      with open(os.path.join(UPLOAD_DIR, fname), "wb") as f: f.write(await file.read())
      image_path = f"/static/uploads/{fname}"
      c.execute("INSERT INTO campaign_popup(image_path,is_active,sort_order,created_at) VALUES(?,?,?,?)",
                (image_path, 1, 0, now))
    con.commit(); con.close()
    return RedirectResponse("/admin/campaigns", status_code=303)

@app.post("/admin/campaign/delete/{cid}")
def admin_campaign_delete(request: Request, cid:int):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("SELECT image_path FROM campaign_popup WHERE id=?", (cid,))
    row = c.fetchone()
    if not row:
        con.close(); raise HTTPException(404, "Kayıt bulunamadı.")
    img = row["image_path"] or ""
    try:
        if img.startswith("/static/uploads/"):
            fs_path = img.lstrip("/")
            full = os.path.abspath(fs_path)
            allowed_root = os.path.abspath(UPLOAD_DIR)
            if full.startswith(allowed_root) and os.path.exists(full):
                os.remove(full)
    except Exception:
        pass
    c.execute("DELETE FROM campaign_popup WHERE id=?", (cid,))
    con.commit(); con.close()
    return RedirectResponse("/admin/campaigns", status_code=303)

# ===================== KULLANICI YÖNETİMİ =====================
@app.get("/admin/users", response_class=HTMLResponse)
def admin_users(request: Request):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("SELECT id, username, is_active, created_at FROM users ORDER BY id ASC")
    users = c.fetchall(); con.close()
    return templates.TemplateResponse("admin_users.html", {
        "request": request, "title": APP_TITLE, "username": request.session.get("user"), "users": users
    })

@app.post("/admin/users/create")
def admin_users_create(request: Request, username: str = Form(...), password: str = Form(...)):
    require_login(request)
    if not username.strip() or not password:
        raise HTTPException(400, "Kullanıcı adı ve şifre zorunludur.")
    con=db(); c=con.cursor()
    try:
        c.execute("INSERT INTO users(username,password,is_active,created_at) VALUES(?,?,1,?)",
                  (username.strip(), password, datetime.utcnow().isoformat(timespec="seconds")))
        con.commit()
    except sqlite3.IntegrityError:
        con.close()
        raise HTTPException(400, "Bu kullanıcı adı zaten mevcut.")
    con.close()
    return RedirectResponse("/admin/users", status_code=303)

@app.post("/admin/users/delete/{uid}")
def admin_users_delete(request: Request, uid:int):
    require_login(request)
    con=db(); c=con.cursor()
    c.execute("SELECT username FROM users WHERE id=?", (uid,))
    row=c.fetchone()
    if not row:
        con.close(); raise HTTPException(404, "Kullanıcı bulunamadı.")
    if row["username"]=="admin":
        con.close(); raise HTTPException(400, "admin kullanıcısı silinemez.")
    c.execute("DELETE FROM users WHERE id=?", (uid,))
    con.commit(); con.close()
    return RedirectResponse("/admin/users", status_code=303)

# ===================== PUBLIC API & BAYİ =====================
class StockItem(BaseModel):
    name: str
    description: str
    list_price: float
    sale_price: float
    cargo_fee: str
    durapay: float
    campaign_categories: List[str]
    product_category: str
    onhand: float
    image_path: str

@app.get("/api/stock", response_model=List[StockItem])
def api_public_stock(search: str = "", category: str = ""):
    con=db(); c=con.cursor()
    params=[CENTER_LOCATION_CODE]
    q="""
    SELECT p.name, p.description, p.list_price, p.sale_price, p.cargo_fee, p.durapay,
           p.campaign_categories, p.product_category,
           COALESCE(ss.onhand,0) as onhand, p.image_path
    FROM product p
    LEFT JOIN location l ON l.code=?
    LEFT JOIN stock_snapshot ss ON ss.product_id=p.id AND ss.location_id=l.id
    WHERE p.is_active=1
    """
    if search:
        q += " AND (p.name LIKE ? OR p.description LIKE ?)"
        like=f"%{search}%"; params.extend([like,like])
    if category and category.lower() != "tümü":
        q += " AND (',' || IFNULL(p.campaign_categories,'') || ',') LIKE ?"
        params.append(f"%,{category},%")
    q += " ORDER BY p.name"
    c.execute(q, tuple(params)); rows=c.fetchall(); con.close()

    items=[]
    for r in rows:
        cc = [x for x in (r["campaign_categories"] or "").strip(",").split(",") if x]
        items.append(StockItem(
            name=r["name"], description=r["description"] or "",
            list_price=float(r["list_price"] or 0), sale_price=float(r["sale_price"] or 0),
            cargo_fee=r["cargo_fee"] or "0", durapay=float(r["durapay"] or 0),
            campaign_categories=cc,
            product_category=r["product_category"] or PRODUCT_CATEGORIES[0],
            onhand=float(r["onhand"] or 0), image_path=r["image_path"] or ""
        ))
    return items

def get_active_campaign_popups() -> List[sqlite3.Row]:
    con=db(); c=con.cursor()
    c.execute("SELECT * FROM campaign_popup WHERE is_active=1 ORDER BY sort_order ASC, id DESC")
    rows=c.fetchall(); con.close()
    return rows

@app.get("/dealer", response_class=HTMLResponse)
def dealer_page(request: Request, search: str = "", category: str = "Tümü"):
    popups = get_active_campaign_popups()
    return templates.TemplateResponse("dealer.html", {
        "request": request, "search": search, "category": category,
        "title": APP_TITLE, "year": datetime.utcnow().year,
        "campaign_cats": CAMPAIGN_CATEGORIES, "popups": popups
    })

# ===================== TEMPLATES =====================
LOGIN_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Giriş</title>
<style>
:root{ --bg:#0f172a; --panel:#0b1227; --line:#1e293b; --text:#e5e7eb; --muted:#94a3b8; }
*{box-sizing:border-box} body{margin:0;min-height:100vh;display:grid;grid-template-rows:1fr auto;background:linear-gradient(180deg,#0b1022,#0f172a);color:var(--text);font:14px/1.5 system-ui,Segoe UI,Arial}
.card{width:min(420px,92vw);background:linear-gradient(180deg,#0f172a,#0b1227);border:1px solid var(--line);border-radius:14px;padding:20px;box-shadow:0 10px 30px rgba(0,0,0,.25);margin:auto}
h1{margin:0 0 12px;font-size:18px} label{display:block;color:var(--muted);margin:10px 0 6px}
input{width:100%;padding:10px 12px;border:1px solid #233143;background:#0b1227;color:#e5e7eb;border-radius:10px;outline:none}
.btn{width:100%;margin-top:14px;padding:10px 14px;border-radius:10px;border:1px solid #1f3b2d;background:linear-gradient(180deg,#06351e,#0f3f2b);color:#bbf7d0;cursor:pointer;white-space:nowrap}
.err{color:#ef4444;font-size:13px;margin-top:8px}.muted{color:#94a3b8;font-size:12px;margin-top:10px;text-align:center}
.footer{padding:16px 0;text-align:center;color:#94a3b8}
</style></head><body>
  <form class="card" method="post" action="/login">
    <h1>Yönetici Girişi</h1>
    <label>Kullanıcı adı</label><input name="username" placeholder="admin" required>
    <label>Şifre</label><input name="password" type="password" placeholder="••••••••" required>
    <button class="btn">Giriş yap</button>
    {% if error %}<div class="err">{{ error }}</div>{% endif %}
    <div class="muted">Bayi görünümü herkese açık: <code>/dealer</code></div>
  </form>
  <div class="footer">2025 • Dijitalizasyon</div>
</body></html>
"""

# --------- ADMIN ORTAK STİL ---------
ADMIN_BASE_STYLE = r"""
:root{ --bg:#0f172a; --line:#1f2937; --text:#e5e7eb; --muted:#94a3b8; }
*{box-sizing:border-box} body{margin:0;min-height:100vh;display:grid;grid-template-rows:auto 1fr auto;background:linear-gradient(180deg,#0b1022,#0f172a);color:#e5e7eb;font:14px/1.5 system-ui,Segoe UI,Arial}
.container{max-width:1600px;margin:auto;padding:16px}
.topbar{position:sticky;top:0;z-index:10;background:rgba(15,23,42,.85);backdrop-filter:blur(8px);border-bottom:1px solid var(--line)}
.topbar .inner{display:flex;gap:12px;align-items:center;justify-content:space-between;padding:12px 16px}
.brand{display:flex;align-items:center;gap:12px}
.logo{height:24px;width:auto;object-fit:contain}
@media(min-width:940px){ .logo{height:28px} }
.title{font-weight:700}
.badge{font-size:12px;padding:4px 10px;border-radius:999px;background:#0b3b2a;color:#a7f3d0;border:1px solid #14532d}
a{color:#38bdf8;text-decoration:none}
.card{background:linear-gradient(180deg,#0f172a,#0b1227); border:1px solid var(--line); border-radius:14px; padding:16px; box-shadow:0 10px 30px rgba(0,0,0,.25); margin-top:16px}
.btn{display:inline-flex;align-items:center;gap:8px;padding:10px 14px;border-radius:10px;border:1px solid #1f3b2d;background:linear-gradient(180deg,#06351e,#0f3f2b);color:#bbf7d0;cursor:pointer;white-space:nowrap}
.badge-draft{font-size:11px;padding:2px 8px;border-radius:999px;background:#3b2a0b;color:#fde68a;border:1px solid #7c5a14}
.badge-live{font-size:11px;padding:2px 8px;border-radius:999px;background:#0b3b2a;color:#a7f3d0;border:1px solid #14532d}
.notice{margin-top:8px;color:#a5b4fc}
.footer{padding:16px 0;text-align:center;color:#94a3b8}
"""

ADMIN_MENU_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Admin Menü</title>
<style>
""" + ADMIN_BASE_STYLE + r"""
.grid{display:grid;gap:16px;grid-template-columns:repeat(auto-fit,minmax(240px,1fr))}
.tile{display:grid;gap:8px;padding:18px;border:1px solid #1f2937;border-radius:14px;background:linear-gradient(180deg,#101935,#0b1227)}
.tile h3{margin:0}
</style></head><body>
  <div class="topbar"><div class="inner container">
    <div class="brand">
      <img class="logo" alt="Logo" src="https://www.google.com/images/branding/googlelogo/2x/googlelogo_color_92x30dp.png">
      <span class="title">Stok Ekranı</span>
      <span class="badge">Admin</span>
    </div>
    <div>Kullanıcı: <strong>{{ username }}</strong> · <a href="/logout">Çıkış</a></div>
  </div></div>

  <div class="container">
    <div class="card">
      <h2>Yönetim Menüsü</h2>
      <div class="grid" style="margin-top:12px">
        <a class="tile" href="/admin/products"><h3>Ürün Yönetimi</h3><div>Ürün ekle, düzenle, stok ve listeyi görüntüle.</div></a>
        <a class="tile" href="/admin/users"><h3>Kullanıcı Yönetimi</h3><div>Yönetici kullanıcıları ekle/sil.</div></a>
        <a class="tile" href="/admin/campaigns"><h3>Kampanya Pop-up Yönetimi</h3><div>Kampanya banner görselleri yükle ve yönet.</div></a>
      </div>
    </div>
  </div>

  <div class="footer">2025 • Dijitalizasyon</div>
</body></html>
"""

ADMIN_PRODUCTS_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Ürün Yönetimi</title>
<style>
""" + ADMIN_BASE_STYLE + r"""
label{display:block;color:#94a3b8;margin:10px 0 6px}
input,textarea,select{width:100%;padding:10px 12px;border:1px solid #233143;background:#0b1227;color:#e5e7eb;border-radius:10px;outline:none}
textarea{min-height:76px}
.table-wrap{border:1px solid #1f2937;border-radius:12px;margin-top:8px;overflow:hidden}
table{width:100%;border-collapse:collapse;table-layout:fixed}
th,td{border-bottom:1px solid #1f2937;padding:12px;text-align:left;vertical-align:middle;word-break:break-word}
thead th{position:sticky;top:0;background:#0c1329}
.thumb{width:56px;height:56px;object-fit:cover;border-radius:8px;border:1px solid #1f2937;background:#0b1227}
.badgecat{display:inline-block;padding:2px 8px;border:1px solid #27406a;border-radius:999px;color:#9ec1ff;background:#0b1630;font-size:12px;margin-right:6px}
.pillset{display:flex;gap:8px;flex-wrap:wrap}
.pill{display:inline-flex;gap:6px;align-items:center;padding:6px 10px;border-radius:999px;border:1px solid #2a3a59;background:#0c1329;color:#e5e7eb;cursor:pointer;font-size:13px}
.pill input{appearance:none;width:14px;height:14px;border:1px solid #47618f;border-radius:4px;display:inline-block;background:transparent;position:relative}
.pill input:checked{background:#3b82f6;border-color:#3b82f6}
.pill input:checked::after{content:"";position:absolute;inset:3px;background:#fff;border-radius:2px}
.price{white-space:nowrap}
.nav{display:flex;gap:10px;margin:10px 0}
.tools{display:grid;gap:10px}
.grid2{display:grid;gap:12px}
@media(min-width:900px){ .grid2{grid-template-columns:1fr 1fr} }
.notice strong{font-weight:700}
.badge-draft{vertical-align:middle}

/* Ürün kodu şeridi — aşağıdan yukarıya ve TAM ORTA */
.codevert{
  writing-mode:vertical-rl;
  text-orientation:mixed;
  letter-spacing:.5px;
  font-weight:700;
  color:#9ec1ff;
  background:#0b1630;
  border:1px solid #27406a;
  border-radius:8px;
  padding:6px 4px;
  display:flex;
  align-items:center;
  justify-content:center;
  min-height:56px; /* admin küçük görsel yüksekliği */
  min-width:20px;
  text-align:center;
  transform:rotate(180deg)
}
</style></head><body>
  <div class="topbar"><div class="inner container">
    <div class="brand">
      <img class="logo" alt="Logo" src="https://www.google.com/images/branding/googlelogo/2x/googlelogo_color_92x30dp.png">
      <span class="title">Stok Ekranı</span>
      <span class="badge">Admin</span>
    </div>
    <div class="nav"><a href="/admin">← Menü</a> · <a href="/admin/users">Kullanıcı Yönetimi</a> · <a href="/admin/campaigns">Kampanya Pop-up</a> · <a href="/logout">Çıkış</a></div>
  </div></div>

  <div class="container">
    <div class="card grid2">
      <div>
        <h3>Ürün Ekle</h3>
        <form method="post" action="/admin/product/create" enctype="multipart/form-data">
          <label>Ürün Görseli</label><input type="file" name="file" accept=".jpg,.jpeg,.png,.webp,.gif">
          <label>Ürün</label><input name="name" required placeholder="Ürün adı / kodu">
          <label>Ürün Detayları</label><textarea name="description" placeholder="Ürün açıklaması..."></textarea>

          <label>Kampanya Kategorisi (çoklu)</label>
          <div class="pillset">
            {% for cat in campaign_cats %}
            <label class="pill"><input type="checkbox" name="campaign_categories" value="{{cat}}"> {{cat}}</label>
            {% endfor %}
          </div>

          <label>Kategori</label>
          <select name="product_category">
            {% for cat in product_cats %}<option value="{{cat}}">{{cat}}</option>{% endfor %}
          </select>

          <label>Liste Fiyatı</label><input name="list_price" type="number" step="0.01" value="0">
          <label>Satış Fiyatı</label><input name="sale_price" type="number" step="0.01" value="0">
          <label>DuraPay (Prim)</label><input name="durapay" type="number" step="0.01" value="0">
          <label>Stok Bilgisi (Adet)</label><input name="stock" type="number" step="1" value="0">
          <label>Kargo Ücreti</label><input name="cargo_fee" type="text" placeholder="Ör. 0, 150, Ücretsiz">

          <div style="display:flex; gap:8px; margin-top:10px">
            <button class="btn" name="save_mode" value="publish">Kaydet</button>
            <button class="btn" name="save_mode" value="draft" style="background:linear-gradient(180deg,#3b2a0b,#4d3a10);border-color:#6b4e12;color:#fde68a">Taslak olarak kaydet</button>
          </div>
        </form>
        <p class="notice">Not: Aynı ürün adı/kodu girilirse eski kayıt silinmez; yenisi otomatik <em>“İsim - 1”</em> şeklinde eklenir.</p>
      </div>

      <div>
        <h3>Excel ile Toplu Yükleme</h3>
        <form method="post" action="/admin/products/upload-excel" enctype="multipart/form-data" class="tools">
          <label>Excel (.xlsx) — Gerekli sütunlar: <strong>Item</strong> ve <strong>Available Qnt</strong></label>
          <input type="file" name="xls" accept=".xlsx" required>
          <button class="btn">Excel'i Yükle</button>
        </form>
        {% if (up_ok or up_new or up_err) %}
          <p class="notice">Son yükleme: <strong>{{ up_ok }}</strong> güncellendi, <strong>{{ up_new }}</strong> yeni taslak, <strong>{{ up_err }}</strong> atlandı.</p>
        {% endif %}
      </div>
    </div>

    <div class="card">
      <h3>Mevcut Ürünler</h3>
      <div class="table-wrap"><table id="admin-table">
        <thead>
          <tr>
            <th>Durum</th>
            <th style="width:28px;text-align:center"></th> <!-- Ürün kodu sütunu -->
            <th>Ürün Görseli</th><th>Ürün</th><th>Ürün Özellikleri</th>
            <th>Kampanya Kategorisi</th><th>Kategori</th>
            <th>Liste Fiyatı</th><th>Satış Fiyatı</th><th>DuraPay</th><th>Stok</th><th>Kargo Ücreti</th><th>Aksiyon</th>
          </tr>
        </thead>
        <tbody>
        {% for p in products %}
          {% set cc = (p.campaign_categories or '').strip(',').split(',') %}
          <tr>
            <td>{% if p.is_active %}<span class="badge-live">Yayında</span>{% else %}<span class="badge-draft">Taslak</span>{% endif %}</td>
            <td style="text-align:center;vertical-align:middle;">
              <div style="display:flex;justify-content:center;">
                <span class="codevert">{{p.name}}</span>
              </div>
            </td>
            <td>{% if p.image_path %}<img class="thumb" src="{{ p.image_path }}">{% else %}<span style="color:#94a3b8">yok</span>{% endif %}</td>
            <td>{{p.name}}</td>
            <td style="white-space:normal">{{p.description}}</td>
            <td>{% for tag in cc if tag %}<span class="badgecat">{{tag}}</span>{% endfor %}</td>
            <td>{{ p.product_category or 'Vitrifiye' }}</td>
            <td><span class="price" data-v="{{p.list_price or 0}}"></span></td>
            <td><span class="price" data-v="{{p.sale_price or 0}}"></span></td>
            <td><span class="price" data-v="{{p.durapay or 0}}"></span></td>
            <td>{{"%.0f"|format(p.onhand or 0)}}</td>
            <td>{{p.cargo_fee}}</td>
            <td><a class="btn" href="/admin/product/edit/{{p.id}}">Düzenle</a></td>
          </tr>
        {% endfor %}
        </tbody></table></div>
    </div>
  </div>

  <div class="footer">2025 • Dijitalizasyon</div>

  <script>
    const fmtTR = (v)=> new Intl.NumberFormat('tr-TR', {maximumFractionDigits:0}).format(Number(v||0)) + ' TL';
    document.querySelectorAll('#admin-table .price').forEach(el=>{ el.textContent = fmtTR(el.getAttribute('data-v')); });
  </script>
</body></html>
"""

EDIT_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Ürün Düzenle</title>
<style>
""" + ADMIN_BASE_STYLE + r"""
label{display:block;color:#94a3b8;margin:10px 0 6px}
input,textarea,select{width:100%;padding:10px 12px;border:1px solid #233143;background:#0b1227;color:#e5e7eb;border-radius:10px;outline:none}
textarea{min-height:76px}
.thumb{width:72px;height:72px;object-fit:cover;border-radius:8px;border:1px solid #1f2937;background:#0b1227}
.badgecat{display:inline-block;padding:2px 8px;border:1px solid #27406a;border-radius:999px;color:#9ec1ff;background:#0b1630;font-size:12px;margin-right:6px}
.pillset{display:flex;gap:8px;flex-wrap:wrap}
.pill{display:inline-flex;gap:6px;align-items:center;padding:6px 10px;border-radius:999px;border:1px solid #2a3a59;background:#0c1329;color:#e5e7eb;cursor:pointer;font-size:13px}
.pill input{appearance:none;width:14px;height:14px;border:1px solid #47618f;border-radius:4px;display:inline-block;background:transparent;position:relative}
.pill input:checked{background:#3b82f6;border-color:#3b82f6}
.pill input:checked::after{content:"";position:absolute;inset:3px;background:#fff;border-radius:2px}
.nav{display:flex;gap:10px;margin:10px 0}
</style></head><body>
  <div class="topbar"><div class="inner container">
    <div class="brand">
      <img class="logo" alt="Logo" src="https://www.google.com/images/branding/googlelogo/2x/googlelogo_color_92x30dp.png">
    <span class="title">Stok Ekranı</span>
      <span class="badge">Admin</span>
    </div>
    <div class="nav"><a href="/admin">← Menü</a> · <a href="/admin/products">Ürün Yönetimi</a> · <a href="/logout">Çıkış</a></div>
  </div></div>

  <div class="container">
    <div class="card">
      <h3>Ürün Düzenle (Yayınla)</h3>
      <form method="post" action="/admin/product/update" enctype="multipart/form-data">
        <input type="hidden" name="pid" value="{{ p.id }}">
        <label>Mevcut Görsel</label>
        {% if p.image_path %}<img class="thumb" src="{{ p.image_path }}">{% else %}<span style="color:#94a3b8">yok</span>{% endif %}

        <label>Yeni Görsel (isteğe bağlı)</label>
        <input type="file" name="file" accept=".jpg,.jpeg,.png,.webp,.gif">

        <label>Ürün</label>
        <input name="name" required value="{{ p.name }}">

        <label>Ürün Detayları</label>
        <textarea name="description">{{ p.description }}</textarea>

        <label>Kampanya Kategorisi (çoklu)</label>
        {% set cc = (p.campaign_categories or '').strip(',').split(',') %}
        <div class="pillset">
          {% for cat in campaign_cats %}
          <label class="pill"><input type="checkbox" name="campaign_categories" value="{{cat}}" {% if cat in cc %}checked{% endif %}> {{cat}}</label>
          {% endfor %}
        </div>

        <label>Kategori</label>
        <select name="product_category">
          {% for cat in product_cats %}
            <option value="{{cat}}" {% if (p.product_category or product_cats[0])==cat %}selected{% endif %}>{{cat}}</option>
          {% endfor %}
        </select>

        <label>Liste Fiyatı</label><input name="list_price" type="number" step="0.01" value="{{ '%.2f'|format(p.list_price or 0) }}">
        <label>Satış Fiyatı</label><input name="sale_price" type="number" step="0.01" value="{{ '%.2f'|format(p.sale_price or 0) }}">
        <label>DuraPay (Prim)</label><input name="durapay" type="number" step="0.01" value="{{ '%.2f'|format(p.durapay or 0) }}">
        <label>Stok Bilgisi (Adet)</label><input name="stock" type="number" step="1" value="{{ '%.0f'|format(stock or 0) }}">
        <label>Kargo Ücreti</label><input name="cargo_fee" type="text" value="{{ p.cargo_fee }}">

        <button class="btn" style="margin-top:10px">Güncelle ve Yayına Al</button>
        <a class="btn" style="margin-top:10px; margin-left:8px" href="/admin/products">İptal</a>
      </form>
    </div>
  </div>

  <div class="footer">2025 • Dijitalizasyon</div>
</body></html>
"""

ADMIN_CAMPAIGNS_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Kampanya Pop-up Yönetimi</title>
<style>
""" + ADMIN_BASE_STYLE + r"""
.grid{display:grid;gap:10px;grid-template-columns:repeat(auto-fill,minmax(180px,1fr))}
.cardimg{position:relative;border:1px solid #1f2937;border-radius:10px;padding:8px;background:linear-gradient(180deg,#0f172a,#0b1227)}
.cardimg img{width:100%;height:120px;object-fit:cover;border-radius:8px;border:1px solid #1f2937;background:#0b1227}
.actions{margin-top:8px;display:flex;gap:8px}
.nav{display:flex;gap:10px;margin:10px 0}
label{display:block;color:#94a3b8;margin:10px 0 6px}
input{width:100%;padding:10px 12px;border:1px solid #233143;background:#0b1227;color:#e5e7eb;border-radius:10px;outline:none}
</style></head><body>
  <div class="topbar"><div class="inner container">
    <div class="brand">
      <img class="logo" alt="Logo" src="https://www.google.com/images/branding/googlelogo/2x/googlelogo_color_92x30dp.png">
      <span class="title">Stok Ekranı</span>
      <span class="badge">Admin</span>
    </div>
    <div class="nav"><a href="/admin">← Menü</a> · <a href="/admin/products">Ürün Yönetimi</a> · <a href="/admin/users">Kullanıcı Yönetimi</a> · <a href="/logout">Çıkış</a></div>
  </div></div>

  <div class="container">
    <div class="card">
      <h3>Kampanya pop-up yükle</h3>
      <form method="post" action="/admin/campaign/upload" enctype="multipart/form-data">
        <label>Görseller (birden fazla seçilebilir)</label>
        <input type="file" name="files" accept=".jpg,.jpeg,.png,.webp,.gif" multiple>
        <button class="btn" style="margin-top:10px">Yükle</button>
      </form>
    </div>

    <div class="card">
      <h3>Yüklü Görseller</h3>
      {% if popups and popups|length > 0 %}
      <div class="grid">
        {% for it in popups %}
          <div class="cardimg">
            <img src="{{ it.image_path }}" alt="kampanya">
            <div class="actions">
              <form method="post" action="/admin/campaign/delete/{{ it.id }}" onsubmit="return confirm('Bu görsel kaldırılsın mı?')">
                <button class="btn" style="background:linear-gradient(180deg,#3f0f0f,#5a1111);border-color:#5f1a1a;color:#fecaca">Kaldır</button>
              </form>
            </div>
          </div>
        {% endfor %}
      </div>
      {% else %}
      <div style="color:#94a3b8">Henüz kampanya görseli yok.</div>
      {% endif %}
    </div>
  </div>

  <div class="footer">2025 • Dijitalizasyon</div>
</body></html>
"""

ADMIN_USERS_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Kullanıcı Yönetimi</title>
<style>
""" + ADMIN_BASE_STYLE + r"""
label{display:block;color:#94a3b8;margin:10px 0 6px}
input{width:100%;padding:10px 12px;border:1px solid #233143;background:#0b1227;color:#e5e7eb;border-radius:10px;outline:none}
.table-wrap{overflow:hidden;border:1px solid #1f2937;border-radius:12px;margin-top:8px}
table{width:100%;border-collapse:collapse;table-layout:fixed}
th,td{border-bottom:1px solid #1f2937;padding:12px;text-align:left;vertical-align:middle;word-break:break-word}
thead th{position:sticky;top:0;background:#0c1329}
.nav{display:flex;gap:10px;margin:10px 0}
</style></head><body>
  <div class="topbar"><div class="inner container">
    <div class="brand">
      <img class="logo" alt="Logo" src="https://www.google.com/images/branding/googlelogo/2x/googlelogo_color_92x30dp.png">
      <span class="title">Stok Ekranı</span>
      <span class="badge">Admin</span>
    </div>
    <div class="nav"><a href="/admin">← Menü</a> · <a href="/admin/products">Ürün Yönetimi</a> · <a href="/admin/campaigns">Kampanya Pop-up</a> · <a href="/logout">Çıkış</a></div>
  </div></div>

  <div class="container">
    <div class="card">
      <h3>Yeni Kullanıcı Ekle</h3>
      <form method="post" action="/admin/users/create">
        <label>Kullanıcı adı</label><input name="username" required>
        <label>Şifre</label><input name="password" type="text" required>
        <button class="btn" style="margin-top:10px">Ekle</button>
      </form>
    </div>

    <div class="card">
      <h3>Mevcut Kullanıcılar</h3>
      <div class="table-wrap"><table>
        <thead><tr><th>ID</th><th>Kullanıcı adı</th><th>Durum</th><th>Oluşturulma</th><th></th></tr></thead>
        <tbody>
          {% for u in users %}
          <tr>
            <td>{{u.id}}</td>
            <td>{{u.username}}</td>
            <td>{{ 'Aktif' if u.is_active else 'Pasif' }}</td>
            <td>{{u.created_at or ''}}</td>
            <td>
              {% if u.username != 'admin' %}
              <form method="post" action="/admin/users/delete/{{u.id}}" onsubmit="return confirm('Silmek istediğinize emin misiniz?')">
                <button class="btn">Sil</button>
              </form>
              {% else %}
                <span style="color:#94a3b8">Silinemez</span>
              {% endif %}
            </td>
          </tr>
          {% endfor %}
        </tbody>
      </table></div>
    </div>
  </div>

  <div class="footer">2025 • Dijitalizasyon</div>
</body></html>
"""

# ------- BAYİ -------
DEALER_HTML = r"""
<!doctype html><html lang="tr"><head>
<meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>{{ title }} · Bayi</title>
<style>
:root{ --line:#1e293b; --text:#e5e7eb; --muted:#93a0b4; }
*{box-sizing:border-box}
body{margin:0;min-height:100vh;display:grid;grid-template-rows:auto 1fr auto;background:linear-gradient(180deg,#091126,#0f172a);color:#e5e7eb;font:14px/1.5 system-ui,Segoe UI,Arial}
.container{max-width:1600px;margin:auto;padding:16px}
.top{position:sticky;top:0;z-index:10;background:rgba(15,23,42,.85);backdrop-filter:blur(8px);border-bottom:1px solid var(--line)}
.top .inner{display:flex;gap:12px;align-items:center;justify-content:space-between;padding:12px 16px}
.brand{display:flex;align-items:center;gap:12px}
.logo{height:28px;width:auto;object-fit:contain}
@media(min-width:940px){ .logo{height:32px} }
.title{font-weight:700}
.badge{font-size:12px;padding:4px 10px;border-radius:999px;background:#0b3b2a;color:#a7f3d0;border:1px solid #14532d}

.bannerrow{display:grid;gap:10px;margin-bottom:12px;grid-template-columns:repeat(auto-fill,minmax(260px,1fr))}
.banner{width:100%;height:120px;object-fit:cover;border-radius:12px;border:1px solid var(--line);background:#0b1227}
@media(min-width:940px){ .banner{height:140px} }

.panel{background:linear-gradient(180deg,#0f172a,#0b1227);border:1px solid var(--line);border-radius:14px;padding:14px;box-shadow:0 10px 30px rgba(0,0,0,.25)}
.controls{display:grid;gap:10px}
@media(min-width:720px){ .controls{grid-template-columns:1fr auto } }
input{width:100%;padding:10px 12px;border:1px solid #22314a;background:#0b1227;color:#e5e7eb;border-radius:10px;outline:none}
.btn{padding:10px 14px;border-radius:10px;border:1px solid #1f3b2d;background:linear-gradient(180deg,#06351e,#0f3f2b);color:#bbf7d0;cursor:pointer;white-space:nowrap}
.muted{color:var(--muted)}

.cats{display:flex;gap:8px;flex-wrap:wrap;margin-top:10px}
.catbtn{padding:6px 10px;border-radius:999px;border:1px solid #2a3a59;background:#0c1329;color:#e5e7eb;cursor:pointer;font-size:13px;white-space:nowrap}
.catbtn.active{border-color:#3b82f6;box-shadow:0 0 0 1px #3b82f6 inset}

/* Seçili kategori başlığı büyük */
.category-heading{
  margin:12px 2px 0;
  font-weight:800;
  font-size:clamp(20px, 3.2vw, 28px);
  letter-spacing:0.2px;
}

.table-wrap{border:1px solid var(--line);border-radius:12px;overflow:hidden}
table{width:100%;border-collapse:collapse;table-layout:fixed}
th,td{padding:12px;border-bottom:1px solid var(--line);text-align:left;vertical-align:middle;word-break:break-word}
thead th{position:sticky;top:0;background:#0c1329}
thead th:nth-child(5), tbody td:nth-child(5){ white-space:normal; max-width:560px } /* Ürün Özellikleri sütunu */
.badgecat{display:inline-block;padding:2px 8px;border:1px solid #27406a;border-radius:999px;color:#9ec1ff;background:#0b1630;font-size:12px;margin-right:6px}

/* Görseller */
.thumb{width:96px;height:96px;object-fit:cover;border-radius:10px;border:1px solid var(--line);background:#0b1227;cursor:pointer}
@media(min-width:940px){ .thumb{width:140px;height:140px} }

/* Ürün kodu şeridi — aşağıdan yukarıya ve TAM ORTA */
.codevert{
  writing-mode:vertical-rl;
  text-orientation:mixed;
  letter-spacing:.5px;
  font-weight:700;
  color:#9ec1ff;
  background:#0b1630;
  border:1px solid #27406a;
  border-radius:8px;
  padding:6px 4px;
  display:flex;
  align-items:center;
  justify-content:center;
  min-height:96px; /* bayi daha büyük görsel ile hizalı */
  min-width:20px;
  text-align:center;
  transform:rotate(180deg)
}

/* DuraPay baloncuk */
.dpwrap{position:relative;display:inline-block}
.dpwrap .bubble{
  position:absolute;left:50%;bottom:100%;transform:translate(-50%,-8px);
  background:#0b3b2a;border:1px solid #14532d;color:#bbf7d0;padding:6px 10px;border-radius:8px;white-space:nowrap;
  opacity:0;pointer-events:none;transition:opacity .15s ease, transform .15s ease;
  box-shadow:0 6px 20px rgba(0,0,0,.35); font-size:13px
}
.dpwrap.open .bubble{opacity:1;pointer-events:auto;transform:translate(-50%,-12px)}
.toggle{padding:6px 10px;border-radius:8px;border:1px solid #1f3b2d;background:linear-gradient(180deg,#06351e,#0f3f2b);color:#bbf7d0;cursor:pointer}

/* Kart (mobil) */
.cards{display:grid;gap:12px}
.card{display:grid;gap:10px;background:linear-gradient(180deg,#0f172a,#0b1227);border:1px solid var(--line);border-radius:12px;padding:12px}
.card-top{display:flex;gap:10px;align-items:stretch}
.kv{display:flex;justify-content:space-between;gap:8px}
.kv .k{color:#93a0b4}
.old-price{text-decoration:line-through;opacity:.8}
.cover{width:100%;object-fit:cover;border-radius:12px;border:1px solid var(--line);background:#0b1227;cursor:pointer;max-height:320px}
@media(min-width:940px){ .cover{max-height:500px} }

/* Masaüstü/mobil geçişi */
@media(min-width:940px){ .cards{display:none} }
@media(max-width:939px){ .table-wrap{display:none} }

/* Lightbox (popup) */
.lightbox{position:fixed;inset:0;background:rgba(0,0,0,.85);display:none;align-items:center;justify-content:center;padding:20px;z-index:50}
.lightbox.open{display:flex}
lightbox img{max-width:92vw;max-height:80vh;border-radius:14px;box-shadow:0 20px 50px rgba(0,0,0,.5)}
.lightbox .close{position:absolute;top:16px;right:16px;background:rgba(255,255,255,.15);border:1px solid rgba(255,255,255,.25);color:#fff;border-radius:999px;padding:8px 12px;cursor:pointer}

/* Footer */
.footer{padding:16px 0;text-align:center;color:#94a3b8}
</style>
</head><body>
  <div class="top">
    <div class="inner container">
      <div class="brand">
        <img class="logo" alt="Logo" src="https://www.google.com/images/branding/googlelogo/2x/googlelogo_color_92x30dp.png">
        <span class="title">Stok Ekranı</span>
      </div>
      <span class="badge">Güncel</span>
    </div>
  </div>

  <div class="container" style="margin-top:16px">

    {% if popups and popups|length > 0 %}
      <div class="bannerrow">
        {% for it in popups %}
          <img class="banner" src="{{ it.image_path }}" alt="kampanya">
        {% endfor %}
      </div>
    {% endif %}

    <div class="panel">
      <form class="controls" method="get">
        <input name="search" value="{{ search }}" placeholder="Ara: Ürün adı veya açıklama">
        <button class="btn">Listele</button>
      </form>

      <div class="cats" id="cats"></div>
      <div id="catHeading" class="category-heading"></div>
    </div>

    <div class="table-wrap" style="margin-top:12px">
      <table id="t">
        <thead>
          <tr>
            <th>Kategori</th>
            <th style="width:28px;text-align:center"></th> <!-- Ürün kodu sütunu -->
            <th>Ürün Görseli</th><th>Ürün</th><th>Ürün Özellikleri</th>
            <th>Kampanya Kategorisi</th>
            <th>Liste Fiyatı</th><th>Satış Fiyatı</th><th>Stok</th><th>Kargo Ücreti</th><th>DuraPay</th>
          </tr>
        </thead>
        <tbody></tbody>
      </table>
    </div>

    <div class="cards" id="cards"></div>
  </div>

  <div id="lightbox" class="lightbox" aria-modal="true" role="dialog">
    <button class="close" aria-label="Kapat">✕</button>
    <img id="lightbox-img" alt="Büyük görsel">
  </div>

  <div class="footer">2025 • Dijitalizasyon</div>

  <script>
    const currentCategory = new URLSearchParams(location.search).get("category") || "Tümü";
    const searchQ = new URLSearchParams(location.search).get("search") || "";
    const api = "/api/stock?" + new URLSearchParams({search: searchQ, category: currentCategory}).toString();

    const CAMPAIGN_LIST = {{ campaign_cats|tojson }};
    const ALL = "Tümü";

    const catsEl = document.getElementById("cats");
    [ALL, ...CAMPAIGN_LIST].forEach(cat=>{
      const b = document.createElement("button");
      b.type = "button";
      b.className = "catbtn" + (cat===currentCategory ? " active" : "");
      b.textContent = cat;
      b.addEventListener("click", ()=>{
        const params = new URLSearchParams(location.search);
        if (cat===ALL) params.delete("category"); else params.set("category", cat);
        location.search = params.toString();
      });
      catsEl.appendChild(b);
    });

    const heading = document.getElementById("catHeading");
    heading.textContent = (currentCategory && currentCategory !== ALL) ? currentCategory : "Tüm Kampanyalar";

    const fmtPriceTR = (v)=> new Intl.NumberFormat('tr-TR', {maximumFractionDigits:0}).format(Number(v||0)) + ' TL';
    const fmtCargo = (v)=>{
      const s = (v ?? '').toString().trim();
      if (s === '') return '';
      const num = Number(s.replace(',', '.'));
      return Number.isFinite(num) ? (new Intl.NumberFormat('tr-TR', {maximumFractionDigits:0}).format(num) + ' TL') : s;
    };

    function openLightbox(src){
      const lb = document.getElementById("lightbox");
      const im = document.getElementById("lightbox-img");
      im.src = src; lb.classList.add("open");
    }
    function closeLightbox(){
      const lb = document.getElementById("lightbox");
      const im = document.getElementById("lightbox-img");
      im.src = ""; lb.classList.remove("open");
    }
    document.getElementById("lightbox").addEventListener("click", closeLightbox);
    document.querySelector("#lightbox .close").addEventListener("click", closeLightbox);
    document.addEventListener("keydown", (e)=>{ if(e.key==="Escape") closeLightbox(); });

    fetch(api).then(r=>r.json()).then(rows=>{
      const tb = document.querySelector("#t tbody");
      const cards = document.getElementById("cards");
      tb.innerHTML = ""; cards.innerHTML = "";

      rows.forEach(r=>{
        const codeCell = `<div style="display:flex;justify-content:center;"><span class="codevert">${r.name}</span></div>`;
        const imgCell = r.image_path
          ? `<img class="thumb" src="${r.image_path}" alt="${r.name}" onclick="openLightbox('${r.image_path}')">`
          : `<span class="muted">yok</span>`;

        const ccHtml = (r.campaign_categories || []).map(tag=>`<span class="badgecat">${tag}</span>`).join(" ");

        const tr = document.createElement("tr");
        const dpId = "dpb_" + Math.random().toString(36).slice(2,8);
        tr.innerHTML = `
          <td><span class="badgecat" style="color:#c7f9ff;background:#0b1f30;border-color:#1d3b5c">${r.product_category || ""}</span></td>
          <td style="text-align:center;vertical-align:middle;">${codeCell}</td>
          <td>${imgCell}</td>
          <td>${r.name}</td>
          <td style="max-width:560px;white-space:normal">${r.description || ""}</td>
          <td>${ccHtml || ""}</td>
          <td><span class="old-price">${fmtPriceTR(r.list_price)}</span></td>
          <td>${fmtPriceTR(r.sale_price)}</td>
          <td>${Number(r.onhand||0).toFixed(0)}</td>
          <td>${fmtCargo(r.cargo_fee)}</td>
          <td>
            <div class="dpwrap" id="${dpId}">
              <div class="bubble">${fmtPriceTR(r.durapay)}</div>
              <button class="toggle" type="button">Görüntüle</button>
            </div>
          </td>`;
        const wrap = tr.querySelector("#"+dpId);
        wrap.querySelector(".toggle").addEventListener("click", ()=>{
          wrap.classList.toggle("open");
          const btn = wrap.querySelector(".toggle");
          btn.textContent = wrap.classList.contains("open") ? "Gizle" : "Görüntüle";
        });
        tb.appendChild(tr);

        // Mobil kart
        const c = document.createElement("div");
        const dpIdM = "dpm_" + Math.random().toString(36).slice(2,8);
        const ccHtmlM = (r.campaign_categories || []).map(tag=>`<span class="badgecat">${tag}</span>`).join(" ");
        c.className = "card";
        c.innerHTML = `
          <div class="card-top">
            <span class="codevert" style="min-height:120px">${r.name}</span>
            ${r.image_path ? `<img class="cover" src="${r.image_path}" alt="${r.name}" onclick="openLightbox('${r.image_path}')">` : ``}
          </div>
          <div class="kv"><span class="k">Kategori</span><span class="v"><span class="badgecat" style="color:#c7f9ff;background:#0b1f30;border-color:#1d3b5c">${r.product_category || ""}</span></span></div>
          <div class="kv"><span class="k">Ürün</span><span class="v">${r.name}</span></div>
          <div class="kv"><span class="k">Kampanya</span><span class="v">${ccHtmlM || "-"}</span></div>
          <div class="kv"><span class="k">Liste Fiyatı</span><span class="v old-price">${fmtPriceTR(r.list_price)}</span></div>
          <div class="kv"><span class="k">Satış Fiyatı</span><span class="v">${fmtPriceTR(r.sale_price)}</span></div>
          <div class="kv"><span class="k">Stok</span><span class="v">${Number(r.onhand||0).toFixed(0)}</span></div>
          <div class="kv"><span class="k">Kargo Ücreti</span><span class="v">${fmtCargo(r.cargo_fee)}</span></div>
          <div class="kv">
            <span class="k">DuraPay</span>
            <span class="v">
              <span class="dpwrap" id="${dpIdM}">
                <div class="bubble">${fmtPriceTR(r.durapay)}</div>
                <button class="toggle" type="button">Görüntüle</button>
              </span>
            </span>
          </div>
          ${r.description ? `<div><span class="k" style="color:#93a0b4">Ürün Özellikleri</span><div style="white-space:normal">${r.description}</div></div>` : ``}
        `;
        const wrapM = c.querySelector("#"+dpIdM);
        wrapM.querySelector(".toggle").addEventListener("click", ()=>{
          wrapM.classList.toggle("open");
          const btn = wrapM.querySelector(".toggle");
          btn.textContent = wrapM.classList.contains("open") ? "Gizle" : "Görüntüle";
        });
        cards.appendChild(c);
      });
    });
  </script>
</body></html>
"""

def _materialize_templates():
    os.makedirs("templates_inline", exist_ok=True)
    with open("templates_inline/login.html","w",encoding="utf-8") as f: f.write(LOGIN_HTML)
    with open("templates_inline/admin_menu.html","w",encoding="utf-8") as f: f.write(ADMIN_MENU_HTML)
    with open("templates_inline/admin_products.html","w",encoding="utf-8") as f: f.write(ADMIN_PRODUCTS_HTML)
    with open("templates_inline/admin_campaigns.html","w",encoding="utf-8") as f: f.write(ADMIN_CAMPAIGNS_HTML)
    with open("templates_inline/admin_users.html","w",encoding="utf-8") as f: f.write(ADMIN_USERS_HTML)
    with open("templates_inline/edit.html","w",encoding="utf-8") as f: f.write(EDIT_HTML)
    with open("templates_inline/dealer.html","w",encoding="utf-8") as f: f.write(DEALER_HTML)
